import openpyxl
import os

#region excel类形式
class excel_class():
    def CreateExcel(self, excelName, isForce=False):
        if os.path.exists(excelName):
            if isForce:
                os.remove(excelName)
            else:
                return True
        wb = openpyxl.Workbook()
        wb.save(excelName)
    def CreateTable(self, excelName, tableNameList, tableDataDic=None):
        '''tableDataDic['表名': [行(int)[列(int)],], [行(int)[列(int)],], ]'''
        wb = openpyxl.load_workbook(excelName)
        for tableName in tableNameList:
            wb.create_sheet(tableName)
            wb.save(excelName)
        for tableName in tableNameList:
            self.WriteTableData(excelName, tableName, tableDataDic and tableDataDic.get(tableName) or False)
        
    def WriteTableData(self, excelName, tableName, tableData):
        '''tableData[行(int)[列(int)], [行(int)[列(int)], ]'''
        if tableData and tableName:
            wb = openpyxl.load_workbook(excelName)
            for hItemIndex in tableData:
                for vItemIndex in tableData[hItemIndex]:
                    wb[tableName].cell(row = hItemIndex+1,column = vItemIndex+1,value = tableData[hItemIndex][vItemIndex])
            wb.save(excelName)
    def ReadTableData(self, excelName, tableName, dataIndexList=None):
        '''
        @param: dataIndexList {行[列(int)]}
        return [行(int)[列(int)]]
        '''
        if not os.path.exists(excelName):
            return None
        wb = openpyxl.load_workbook(excelName)
        sheet = wb[tableName]
        maxH = sheet.max_row      #获取最大行数
        maxV = sheet.max_column      #获取最大列数
        list = {}
        if not dataIndexList:
            for itemH in range(maxH):
                list1 = {}
                for itemV in range(maxV):
                    list1[itemV] = sheet.cell(row = itemH+1,column = itemV+1).value
                list[itemH] = list1
        else:
            for itemH in dataIndexList:
                list1 = {}
                for itemV in dataIndexList[itemH]:
                    list1[itemV] = sheet.cell(row=itemH+1, column=itemV+1).value
                list[itemH] = list1

        return list if len(list)!=0 else None
    def ReadTableData_By_ID(self, excelName, tableName, keyIndex, value, dataIndexList=None):
        dataAll = self.ReadTableData(excelName, tableName, dataIndexList)
        for item in dataAll:
            if dataAll[item][keyIndex] == value:
                return dataAll[item]
        return None
    def ChangeTableData(self, excelName, tableName, tableDataDic):
        '''tableDataDic {行{列}}'''
        if not os.path.exists(excelName):
            return None
        wb = openpyxl.load_workbook(excelName)
        sheet = wb[tableName]
        for itemH in tableDataDic:
            for itemV in tableDataDic[itemH]:
                sheet.cell(row=itemH+1, column=itemV+1, value=tableDataDic[itemH][itemV])
        wb.save(excelName)
    def GetMaxHV(self, excelName, tableName):
        '''
        @param: dataIndexList {行[列(int)]}
        return [行(int)[列(int)]]
        '''
        if not os.path.exists(excelName):
            return None
        wb = openpyxl.load_workbook(excelName)
        sheet = wb[tableName]
        maxH = sheet.max_row      #获取最大行数
        maxV = sheet.max_column      #获取最大列数
        return (maxH, maxV)
#endregion

#region excel单体形式
def excel_create(excelName, isForce=False):
    if os.path.exists(excelName):
        if isForce:
            os.remove(excelName)
        else:
            return True
    wb = openpyxl.Workbook()
    wb.save(excelName)
def excel_create_table(excelName, tableNameList, tableDataDic=None, isExitsTabDel=False):
    '''tableDataDic['表名': [行(int)[列(int)],], [行(int)[列(int)],], ]'''
    wb = openpyxl.load_workbook(excelName)
    for tableName in tableNameList:
        if tableName in wb:
            if isExitsTabDel:
                del wb[tableName]
                wb.create_sheet(tableName)
        else:
            wb.create_sheet(tableName)
        wb.save(excelName)
    for tableName in tableNameList:
        excel_write_table(excelName, tableName, tableDataDic and tableDataDic.get(tableName) or False)
def excel_write_table(excelName, tableName, tableData):
    '''tableData[行(int)[列(int)], [行(int)[列(int)], ]'''
    if tableData and tableName:
        wb = openpyxl.load_workbook(excelName)
        for hItemIndex in tableData:
            for vItemIndex in tableData[hItemIndex]:
                wb[tableName].cell(row = hItemIndex+1,column = vItemIndex+1,value = tableData[hItemIndex][vItemIndex])
        wb.save(excelName)
def excel_read_table(excelName, tableName, dataIndexList=None):
    '''
    @param: dataIndexList {行[列(int)]}
    return [行(int)[列(int)]]
    '''
    if not os.path.exists(excelName):
        return None
    wb = openpyxl.load_workbook(excelName)
    sheet = wb[tableName]
    maxH = sheet.max_row      #获取最大行数
    maxV = sheet.max_column      #获取最大列数
    list = {}
    if not dataIndexList:
        for itemH in range(maxH):
            list1 = {}
            for itemV in range(maxV):
                list1[itemV] = sheet.cell(row = itemH+1,column = itemV+1).value
            list[itemH] = list1
    else:
        for itemH in dataIndexList:
            list1 = {}
            for itemV in dataIndexList[itemH]:
                list1[itemV] = sheet.cell(row=itemH+1, column=itemV+1).value
            list[itemH] = list1

    return list if len(list)!=0 else None
def excel_read_table_by_id(excelName, tableName, keyIndex, value, dataIndexList=None):
    dataAll = excel_read_table(excelName, tableName, dataIndexList)
    for item in dataAll:
        if dataAll[item][keyIndex] == value:
            return dataAll[item]
    return None
def excel_change_table(excelName, tableName, tableDataDic):
    '''tableDataDic {行{列}}'''
    if not os.path.exists(excelName):
        return None
    wb = openpyxl.load_workbook(excelName)
    sheet = wb[tableName]
    for itemH in tableDataDic:
        for itemV in tableDataDic[itemH]:
            sheet.cell(row=itemH+1, column=itemV+1, value=tableDataDic[itemH][itemV])
    wb.save(excelName)
def excel_get_max_hv(excelName, tableName):
    '''
    @param: dataIndexList {行[列(int)]}
    return [行(int)[列(int)]]
    '''
    if not os.path.exists(excelName):
        return None
    wb = openpyxl.load_workbook(excelName)
    sheet = wb[tableName]
    maxH = sheet.max_row      #获取最大行数
    maxV = sheet.max_column      #获取最大列数
    return (maxH, maxV)
#endregion

if __name__ == "__main__":
    excelService = excel_class()
    excelService.CreateExcel('./Test.xlsx', True)
    excelService.CreateTable('./Test.xlsx', ['Sheel1','TestT1'], {
        'Sheel1':{
            0:{0:5,1:'xx',2:'678'},
        },
        'TestT1':{
            0:{0:343,1:'xx',2:'678'},
            1:{0:343,1:'TestT1xx',2:'678TestT1'},
        }
    })
    print(excelService.ReadTableData('./Test.xlsx', 'TestT1'))
    print(excelService.ReadTableData('./Test.xlsx', 'TestT1', {0:[0,2],1:[1,2]}))
    excelService.ChangeTableData('./Test.xlsx', 'TestT1',{0:{0:'996',1:'一个大西瓜'}})